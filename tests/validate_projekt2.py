"""Validate parser output against Manuel's reference Auswertung.xlsx for Projekt 2."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import numpy as np
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from mpc2_parser.core import process_measurement  # noqa: E402


PROJEKT_2 = Path(
    "/Users/malcolmwerchota/Library/CloudStorage/OneDrive-SharedLibraries-MPC²/"
    "Manuel Prohaska - MPC²_Beratungsprojekt Green Startupmark/"
    "Messdateien/Projekt 2"
)
AUSWERTUNG = PROJEKT_2 / "Auswertung_ON2025-0003_UNS S32906 3D III_DL-EPR-Test.xlsx"


def extract_manual(xlsx: Path) -> dict[str, dict]:
    wb_f = openpyxl.load_workbook(xlsx, data_only=False)
    wb_v = openpyxl.load_workbook(xlsx, data_only=True)
    out = {}
    for name in wb_f.sheetnames:
        ws_f, ws_v = wb_f[name], wb_v[name]
        if ws_f["L2"].value != "Ja":
            continue
        jr_formula = str(ws_f["M3"].value or "")
        m = re.search(r"F(\d+):F(\d+)", jr_formula)
        out[name] = {
            "ja": ws_v["M2"].value,
            "jr": ws_v["M3"].value,
            "qa": ws_v["M5"].value,
            "qr": ws_v["M6"].value,
            "ruhe": ws_v["P6"].value,
            "split_excel_row": int(m.group(1)) if m else None,
            "split_index": int(m.group(1)) - 2 if m else None,
        }
    return out


def match_asc(sheet: str, asc_files: list[Path]) -> Path | None:
    m = re.match(r"^(\d+)_", sheet)
    if not m:
        return None
    num = m.group(1).lstrip("0") or "0"
    tail = sheet.split("_")[-1]
    for asc in asc_files:
        n = re.match(r"^0*(\d+)_", asc.name)
        if n and n.group(1) == num and tail in asc.name:
            return asc
    for asc in asc_files:
        n = re.match(r"^0*(\d+)_", asc.name)
        if n and n.group(1) == num:
            return asc
    return None


def pct_err(ours, manual):
    if not isinstance(manual, (int, float)) or manual == 0:
        return None
    return 100.0 * (ours - manual) / abs(manual)


def main():
    ref = extract_manual(AUSWERTUNG)
    ascs = sorted(PROJEKT_2.glob("*.ASC"))
    print(f"Auswertung sheets: {len(ref)} | ASC files: {len(ascs)}\n")

    errs = {"ja": [], "jr": [], "qa": [], "qr": []}
    errs_mid = {"ja": [], "jr": [], "qa": [], "qr": []}

    for sheet, r in ref.items():
        asc = match_asc(sheet, ascs)
        if not asc:
            print(f"[SKIP] {sheet}: no matching ASC")
            continue

        try:
            v = process_measurement(asc, split_method="vertex").analysis
            m = process_measurement(asc, split_method="midpoint").analysis
        except Exception as e:
            print(f"[ERROR] {sheet}: {e}")
            continue

        print(f"--- {sheet}  (ASC: {asc.name[:50]}...)")
        print(f"    Manual split row: {r['split_excel_row']} (array idx {r['split_index']})")
        print(f"    Vertex   split idx: {v.split_index}    Midpoint split idx: {m.split_index}")

        def fmt(label, ours_v, ours_m, manual):
            pe_v = pct_err(ours_v, manual)
            pe_m = pct_err(ours_m, manual)
            mv = f"{manual:.6f}" if isinstance(manual, (int, float)) else "---"
            print(f"    {label:6} manual={mv}  vertex={ours_v:.6f} ({pe_v:+.2f}%)  "
                  f"midpoint={ours_m:.6f} ({pe_m:+.2f}%)")
            return pe_v, pe_m

        for key, ours_v, ours_m in [
            ("Ja", v.ja_ma_cm2, m.ja_ma_cm2),
            ("Jr", v.jr_ma_cm2, m.jr_ma_cm2),
            ("Qa", v.qa_as, m.qa_as),
            ("Qr", v.qr_as, m.qr_as),
        ]:
            ev, em = fmt(key, ours_v, ours_m, r[key.lower()])
            if ev is not None:
                errs[key.lower()].append(abs(ev))
            if em is not None:
                errs_mid[key.lower()].append(abs(em))
        print()

    print("\n=== Summary (|% error| vs. Manuel) ===")
    print(f"{'Metric':<6} {'VERTEX mean':<14} {'VERTEX max':<14} {'MIDPOINT mean':<16} {'MIDPOINT max':<14}")
    for k in ("ja", "jr", "qa", "qr"):
        v_arr = np.array(errs[k]) if errs[k] else np.array([0.0])
        m_arr = np.array(errs_mid[k]) if errs_mid[k] else np.array([0.0])
        print(f"{k:<6} {v_arr.mean():<14.2f} {v_arr.max():<14.2f} {m_arr.mean():<16.2f} {m_arr.max():<14.2f}")


if __name__ == "__main__":
    main()
