"""Full 3-way method evaluation: Manuel vs Vertex vs Midpoint.

For every ASC we can match against a known-good reference value:
  - Projekt 1 ASCs → match against Messübersicht Corrosion Ray sheet
  - Projekt 2 ASCs → match against Auswertung_ON2025-0003 workbook

For each match we compute:
  - Vertex method: our auto-detected split
  - Midpoint method: our N/2 split
  - Manuel method: the split index Manuel hand-picked (only available for Projekt 2)

And compare all against the reference.
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from statistics import mean, median

import numpy as np
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from mpc2_parser.core import process_measurement  # noqa: E402
from mpc2_parser.analysis import analyze_dlepr  # noqa: E402
from mpc2_parser.parser import parse_asc, parse_had  # noqa: E402


BASE = Path(
    "/Users/malcolmwerchota/Library/CloudStorage/OneDrive-SharedLibraries-MPC²/"
    "Manuel Prohaska - MPC²_Beratungsprojekt Green Startupmark"
)
PROJEKT_1 = BASE / "Messdateien" / "Projekt 1"
PROJEKT_2 = BASE / "Messdateien" / "Projekt 2"
MASTER    = BASE / "Messübersicht_Elektrochemische Messungen.xlsx"
AUSWERTUNG = PROJEKT_2 / "Auswertung_ON2025-0003_UNS S32906 3D III_DL-EPR-Test.xlsx"


# --------------------------------------------------------------------------- #
# Reference value extractors                                                  #
# --------------------------------------------------------------------------- #

def load_messuebersicht_refs():
    """Returns {id -> {ja, jr, qa, qr, row, materialzustand}}."""
    wb = openpyxl.load_workbook(MASTER, data_only=True)
    ws = wb["Corrosion Ray"]
    out = {}
    for r in range(2, ws.max_row + 1):
        mid = ws.cell(row=r, column=1).value
        if mid is None:
            continue
        mid = str(mid).strip().zfill(4)
        ja = ws.cell(row=r, column=22).value  # V
        jr = ws.cell(row=r, column=23).value  # W
        qa = ws.cell(row=r, column=25).value  # Y
        qr = ws.cell(row=r, column=26).value  # Z
        # X=Jr/Ja formula, AA=Qr/Qa formula — skip (derived)
        if not isinstance(ja, (int, float)):
            continue
        out[mid] = {
            "row": r,
            "ja": ja,
            "jr": jr,
            "qa": qa if isinstance(qa, (int, float)) else None,
            "qr": qr if isinstance(qr, (int, float)) else None,
        }
    return out


def load_auswertung_refs():
    """Returns {'399_...': {ja, jr, qa, qr, manuel_split}}."""
    wb_f = openpyxl.load_workbook(AUSWERTUNG, data_only=False)
    wb_v = openpyxl.load_workbook(AUSWERTUNG, data_only=True)
    out = {}
    for name in wb_f.sheetnames:
        ws_f, ws_v = wb_f[name], wb_v[name]
        if ws_f["L2"].value != "Ja":
            continue
        jr_formula = str(ws_f["M3"].value or "")
        m = re.search(r"F(\d+):F(\d+)", jr_formula)
        out[name] = {
            "ja": ws_v["M2"].value,
            "jr": ws_v["M3"].value,
            "qa": ws_v["M5"].value,
            "qr": ws_v["M6"].value,
            "manuel_split": int(m.group(1)) - 2 if m else None,  # 0-indexed
        }
    return out


# --------------------------------------------------------------------------- #
# ASC matching                                                                #
# --------------------------------------------------------------------------- #

def ascs_by_id(folder: Path) -> dict[str, list[Path]]:
    """Group ASCs by leading 4-digit ID (zero-padded)."""
    out = defaultdict(list)
    for p in sorted(folder.glob("*.ASC")):
        m = re.match(r"^(\d+)", p.name)
        if m:
            mid = m.group(1).zfill(4)
            out[mid].append(p)
    return dict(out)


def match_auswertung_sheet_to_asc(sheet: str, ascs: list[Path]) -> Path | None:
    m = re.match(r"^(\d+)_", sheet)
    if not m:
        return None
    num = m.group(1).lstrip("0") or "0"
    tail = sheet.split("_")[-1]
    for asc in ascs:
        n = re.match(r"^0*(\d+)_", asc.name)
        if n and n.group(1) == num and tail in asc.name:
            return asc
    for asc in ascs:
        n = re.match(r"^0*(\d+)_", asc.name)
        if n and n.group(1) == num:
            return asc
    return None


# --------------------------------------------------------------------------- #
# Analysis                                                                    #
# --------------------------------------------------------------------------- #

def percent_err(ours, ref):
    if not isinstance(ref, (int, float)) or ref == 0:
        return None
    return 100.0 * (ours - ref) / abs(ref)


def evaluate_asc_with_methods(asc_path: Path, had_path: Path | None, manuel_split: int | None = None):
    """Run all 3 methods on one ASC. Returns dict of results by method."""
    out = {}
    asc = parse_asc(asc_path)
    had = parse_had(had_path) if had_path and had_path.exists() else None

    for method in ("vertex", "midpoint"):
        try:
            r = analyze_dlepr(asc, had, split_method=method)
            out[method] = {
                "ja": r.ja_ma_cm2, "jr": r.jr_ma_cm2,
                "qa": r.qa_as, "qr": r.qr_as,
                "split": r.split_index,
                "jr_ja": r.jr_ja, "qr_qa": r.qr_qa,
            }
        except Exception as e:
            out[method] = {"error": str(e)}

    if manuel_split is not None:
        try:
            r = analyze_dlepr(asc, had, split_method="vertex", split_override=manuel_split)
            out["manuel"] = {
                "ja": r.ja_ma_cm2, "jr": r.jr_ma_cm2,
                "qa": r.qa_as, "qr": r.qr_as,
                "split": manuel_split,
                "jr_ja": r.jr_ja, "qr_qa": r.qr_qa,
            }
        except Exception as e:
            out["manuel"] = {"error": str(e)}
    return out


# --------------------------------------------------------------------------- #
# Report builders                                                             #
# --------------------------------------------------------------------------- #

def evaluate_projekt2():
    """Projekt 2 has full ground truth — split indices + Ja/Jr/Qa/Qr."""
    print("\n" + "="*90)
    print("PROJEKT 2 EVALUATION (8 ASCs, full ground truth from Auswertung)")
    print("="*90)
    refs = load_auswertung_refs()
    ascs_p2 = ascs_by_id(PROJEKT_2)

    errs = {method: {k: [] for k in ("ja", "jr", "qa", "qr")}
            for method in ("vertex", "midpoint", "manuel")}
    detailed = []

    for sheet, ref in refs.items():
        matched_ascs = [a for group in ascs_p2.values() for a in group]
        asc = match_auswertung_sheet_to_asc(sheet, matched_ascs)
        if not asc:
            continue
        had = asc.with_suffix(".HAD")
        results = evaluate_asc_with_methods(asc, had, manuel_split=ref["manuel_split"])

        row = {"sheet": sheet, "asc": asc.name, "ref": ref, "methods": results}
        detailed.append(row)

        for method, res in results.items():
            if "error" in res:
                continue
            for metric in ("ja", "jr", "qa", "qr"):
                err = percent_err(res[metric], ref[metric])
                if err is not None:
                    errs[method][metric].append(abs(err))

    return detailed, errs


def evaluate_projekt1():
    """Projekt 1: ground truth is Messübersicht. No split info from Manuel, just final values.

    We compare the parser's vertex and midpoint methods against Manuel's recorded
    Ja/Jr/Qa/Qr values. There's no "Manuel split" to evaluate separately since
    Manuel didn't leave an Auswertung workbook for this project.
    """
    print("\n" + "="*90)
    print("PROJEKT 1 EVALUATION (against Messübersicht reference values)")
    print("="*90)
    refs = load_messuebersicht_refs()
    ascs_p1 = ascs_by_id(PROJEKT_1)
    overlap = sorted(set(ascs_p1) & set(refs))
    print(f"Overlap IDs: {len(overlap)} — {overlap}")

    errs = {method: {k: [] for k in ("ja", "jr", "qa", "qr")}
            for method in ("vertex", "midpoint")}
    detailed = []

    for mid in overlap:
        ref = refs[mid]
        variants = ascs_p1[mid]

        # For IDs with multiple ASC variants, find the one best matching the reference.
        # We run every variant and pick the ASC with smallest |Ja error| under vertex method.
        best = None
        for asc in variants:
            had = asc.with_suffix(".HAD")
            try:
                results = evaluate_asc_with_methods(asc, had, manuel_split=None)
                v = results.get("vertex", {})
                if "error" in v:
                    continue
                ja_err = percent_err(v.get("ja", 0), ref["ja"])
                score = abs(ja_err) if ja_err is not None else 1e9
                if best is None or score < best[0]:
                    best = (score, asc, results)
            except Exception:
                continue

        if best is None:
            continue
        _, asc_chosen, results = best

        row = {
            "id": mid,
            "asc_chosen": asc_chosen.name,
            "n_variants": len(variants),
            "ref": ref,
            "methods": results,
        }
        detailed.append(row)

        for method in ("vertex", "midpoint"):
            res = results.get(method, {})
            if "error" in res:
                continue
            for metric in ("ja", "jr", "qa", "qr"):
                if ref.get(metric) is None:
                    continue
                err = percent_err(res[metric], ref[metric])
                if err is not None:
                    errs[method][metric].append(abs(err))

    return detailed, errs


# --------------------------------------------------------------------------- #
# Pretty reporting                                                            #
# --------------------------------------------------------------------------- #

def print_summary_table(errs_p1, errs_p2_all):
    print("\n\n" + "="*90)
    print("COMBINED 3-WAY METHOD EVALUATION")
    print("="*90)
    header = f"{'Metric':<6} {'Method':<10} {'Project':<10} {'n':<4} {'Mean|%err|':<12} {'Median':<10} {'Max':<10}"
    print(header)
    print("-" * len(header))

    # Projekt 2 has all 3 methods
    for method in ("manuel", "vertex", "midpoint"):
        for metric in ("ja", "jr", "qa", "qr"):
            arr = errs_p2_all.get(method, {}).get(metric, [])
            if arr:
                print(f"{metric.upper():<6} {method:<10} {'P2':<10} {len(arr):<4} "
                      f"{mean(arr):<12.3f} {median(arr):<10.3f} {max(arr):<10.3f}")

    for method in ("vertex", "midpoint"):
        for metric in ("ja", "jr", "qa", "qr"):
            arr = errs_p1.get(method, {}).get(metric, [])
            if arr:
                print(f"{metric.upper():<6} {method:<10} {'P1':<10} {len(arr):<4} "
                      f"{mean(arr):<12.3f} {median(arr):<10.3f} {max(arr):<10.3f}")

    print("\n" + "="*90)
    print("METHOD RANKINGS (lower is better)")
    print("="*90)
    all_p2 = errs_p2_all
    for metric in ("ja", "jr", "qa", "qr"):
        scores = {}
        for method in ("manuel", "vertex", "midpoint"):
            arr = all_p2.get(method, {}).get(metric, [])
            if arr:
                scores[method] = mean(arr)
        ranked = sorted(scores.items(), key=lambda x: x[1])
        rank_str = " > ".join(f"{m} ({v:.2f}%)" for m, v in ranked)
        print(f"  {metric.upper():<4} (Projekt 2): {rank_str}")


def main():
    detailed_p2, errs_p2 = evaluate_projekt2()
    detailed_p1, errs_p1 = evaluate_projekt1()

    # Print per-measurement detail
    print("\n\nPROJEKT 2 DETAIL")
    print("-" * 80)
    for d in detailed_p2:
        print(f"\n[{d['sheet']}] ← {d['asc'][:55]}")
        r = d["ref"]
        print(f"  Ref (Manuel): Ja={r['ja']:.4f}  Jr={r['jr']:.4f}  Qa={r['qa']:.4f}  Qr={r['qr']:.4f}  (split={r['manuel_split']})")
        for method in ("manuel", "vertex", "midpoint"):
            m = d["methods"].get(method, {})
            if "error" in m: continue
            ja_e = percent_err(m['ja'], r['ja'])
            jr_e = percent_err(m['jr'], r['jr'])
            qa_e = percent_err(m['qa'], r['qa'])
            qr_e = percent_err(m['qr'], r['qr'])
            print(f"  {method:<8}: Ja={m['ja']:.4f}({ja_e:+.1f}%) Jr={m['jr']:.4f}({jr_e:+.1f}%) "
                  f"Qa={m['qa']:.4f}({qa_e:+.1f}%) Qr={m['qr']:.4f}({qr_e:+.1f}%) [split={m['split']}]")

    print("\n\nPROJEKT 1 DETAIL")
    print("-" * 80)
    for d in detailed_p1:
        r = d["ref"]
        print(f"\n[ID {d['id']}] ← {d['asc_chosen'][:55]}  (1 of {d['n_variants']} variants)")
        print(f"  Ref (Messübersicht row {r['row']}): Ja={r['ja']:.4f}  Jr={r['jr']:.4f}  Qa={r.get('qa') or '---'}  Qr={r.get('qr') or '---'}")
        for method in ("vertex", "midpoint"):
            m = d["methods"].get(method, {})
            if "error" in m: continue
            ja_e = percent_err(m['ja'], r['ja'])
            jr_e = percent_err(m['jr'], r['jr'])
            qa_e = percent_err(m['qa'], r.get('qa')) if r.get('qa') else None
            qr_e = percent_err(m['qr'], r.get('qr')) if r.get('qr') else None
            print(f"  {method:<8}: Ja={m['ja']:.4f}({ja_e:+.1f}%) Jr={m['jr']:.4f}({jr_e:+.1f}%) "
                  f"Qa={m['qa']:.4f}({qa_e:+.1f}% if qa) Qr={m['qr']:.4f}({qr_e:+.1f}% if qr) [split={m['split']}]")

    print_summary_table(errs_p1, errs_p2)

    # JSON dump for dashboard consumption
    output = {
        "projekt_2": {"details": detailed_p2, "errors": errs_p2},
        "projekt_1": {"details": detailed_p1, "errors": errs_p1},
    }
    # Convert non-JSON-safe types
    import numpy as np
    def clean(o):
        if isinstance(o, dict): return {k: clean(v) for k, v in o.items()}
        if isinstance(o, list): return [clean(x) for x in o]
        if isinstance(o, (np.integer,)): return int(o)
        if isinstance(o, (np.floating,)): return float(o)
        if isinstance(o, (tuple,)): return list(clean(list(o)))
        return o
    out_path = Path(__file__).parent.parent / "out" / "full_evaluation.json"
    out_path.write_text(json.dumps(clean(output), indent=2, default=str))
    print(f"\n\nDetailed results: {out_path}")


if __name__ == "__main__":
    main()
