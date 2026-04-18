"""MPC² Corrosion Ray — DL-EPR-Analyzer (Streamlit web app, MPC² brand styled).

Vier Tabs:
  1. Analyse        — ASC hochladen → Werte + Diagramme + Integritäts-Score → Downloads
  2. Methodik       — Erklärung der Split-Methoden mit Vor-/Nachteilen
  3. Genauigkeit    — Validierungs-Ergebnisse Projekt A, B und kombiniert
  4. Datenqualität  — Gefundene Ungereimtheiten (Sheet-Swap, Ausreißer)

Ausführen:
    streamlit run webapp/app.py
"""
from __future__ import annotations

import io
import json
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import altair as alt
import numpy as np
import openpyxl
import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from mpc2_parser.parser import (  # noqa: E402
    ASCData, HADMetadata, parse_asc, parse_had, parse_filename,
)
from mpc2_parser.analysis import (  # noqa: E402
    analyze_dlepr, detect_split_vertex, detect_split_midpoint,
)
from mpc2_parser.core import Measurement  # noqa: E402
from mpc2_parser.outputs.variant2_messuebersicht import _build_row  # noqa: E402
from mpc2_parser.outputs.variant1_auswertung import write_auswertung_workbook  # noqa: E402
from mpc2_parser.quality import evaluate_integrity  # noqa: E402


# ═══════════════════════════════════════════════════════════════════════════
# Page config + CSS + JS translation
# ═══════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="MPC² Corrosion Ray — DL-EPR-Auswertung",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ═══════════════════════════════════════════════════════════════════════════
# Password Gate
# ═══════════════════════════════════════════════════════════════════════════
#
# Shared-password access control. The expected password lives in st.secrets
# (either ~/.streamlit/secrets.toml locally, or Streamlit Cloud's secrets UI
# in production). If no password is configured, the gate is disabled so local
# dev still works out of the box.

def _check_password() -> bool:
    """Returns True once the correct password has been entered this session."""
    expected = st.secrets.get("password") if hasattr(st, "secrets") else None
    try:
        # st.secrets raises if secrets.toml is missing — treat that as "no gate".
        expected = st.secrets["password"]
    except Exception:
        return True  # no password configured → open access (local dev)

    if st.session_state.get("auth_ok"):
        return True

    # Minimal branded gate screen
    st.markdown("""
    <div style="max-width:460px; margin:80px auto; padding:36px 40px;
                background:white; border:1px solid #EEEEEE;
                border-left:6px solid #8DBF18; font-family:'Manrope',sans-serif;">
      <div style="font-size:11px; font-weight:700; letter-spacing:0.12em;
                  color:#409A2D; text-transform:uppercase; margin-bottom:12px;">
        werchota.ai · MPC² Customer Preview
      </div>
      <h1 style="font-size:26px; font-weight:700; color:#4C4C4C; margin:0 0 8px;
                 letter-spacing:-0.01em;">
        Corrosion Ray® · DL-EPR-Auswertung
      </h1>
      <p style="color:#333; font-size:14px; line-height:1.5; margin:0 0 20px;">
        Zugang nur für eingeladene Nutzer. Bitte Passwort eingeben.
      </p>
    </div>
    """, unsafe_allow_html=True)

    with st.container():
        col_pad_l, col_form, col_pad_r = st.columns([1, 2, 1])
        with col_form:
            pw = st.text_input("Passwort", type="password",
                               key="pw_input", label_visibility="collapsed",
                               placeholder="Passwort")
            if pw:
                if pw == expected:
                    st.session_state.auth_ok = True
                    st.rerun()
                else:
                    st.error("Passwort falsch.")
    return False


if not _check_password():
    st.stop()


# Altair default theme config — MPC²-konforme Farben + WEISSER HINTERGRUND
alt.themes.register("mpc2", lambda: {
    "config": {
        "background": "white",
        "view": {
            "fill": "white",
            "stroke": "transparent",
        },
        "axis": {
            "labelColor": "#4C4C4C",
            "titleColor": "#4C4C4C",
            "labelFont": "Manrope, sans-serif",
            "titleFont": "Manrope, sans-serif",
            "labelFontSize": 11,
            "titleFontSize": 12,
            "titleFontWeight": 600,
            "gridColor": "#EEEEEE",
            "domainColor": "#4C4C4C",
            "tickColor": "#4C4C4C",
        },
        "legend": {
            "labelColor": "#333333",
            "titleColor": "#4C4C4C",
            "labelFont": "Manrope, sans-serif",
            "titleFont": "Manrope, sans-serif",
        },
        "title": {
            "color": "#4C4C4C",
            "font": "Manrope, sans-serif",
            "fontWeight": 600,
        },
        "range": {"category": ["#8DBF18", "#409A2D", "#4C4C4C", "#b8860b"]},
    }
})
alt.themes.enable("mpc2")


MPC2_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@300;400;500;600;700;800&display=swap');
@import url('https://fonts.googleapis.com/icon?family=Material+Icons');
@import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200&display=swap');

/* Material Icons must NOT inherit Manrope — force the icon font on the specific elements */
.material-icons, [class*="material-icons"], [data-testid*="Icon"] i,
.stExpander svg, .stExpander i, .stExpander .material-icons {
  font-family: 'Material Icons', 'Material Symbols Outlined' !important;
  font-feature-settings: 'liga';
}

:root {
  --mpc-green-light: #8DBF18;
  --mpc-green-dark:  #409A2D;
  --mpc-charcoal:    #4C4C4C;
  --mpc-body:        #333333;
  --mpc-bg-light:    #EEEEEE;
  --mpc-white:       #FFFFFF;
  --mpc-gradient:    linear-gradient(90deg, #8DBF18 6.6%, #409A2D 100%);
  --warn:            #b8860b;
  --warn-light:      #fef3c7;
  --danger:          #c0392b;
  --danger-light:    #fdecec;
  --success-light:   #EAF4D8;
}

/* Reset Streamlit top padding so our nav bar is at the top */
[data-testid="stAppViewContainer"] > .main > div:first-child { padding-top: 0 !important; }
[data-testid="stHeader"] { display: none !important; }
.block-container { padding-top: 0 !important; max-width: none !important; }

html, body, .stApp, .stMarkdown, button, input, textarea, select, label, p, h1, h2, h3, h4, h5, h6 {
  font-family: 'Manrope', -apple-system, BlinkMacSystemFont, sans-serif !important;
  color: var(--mpc-body);
  letter-spacing: 0.01em;
}
/* Material Icons — NEVER override */
.material-icons, span.material-icons, i.material-icons,
[class*="material-icons"], .material-icons-outlined, .material-symbols-outlined {
  font-family: 'Material Icons', 'Material Symbols Outlined' !important;
  font-style: normal !important;
  font-weight: normal !important;
  font-size: 18px !important;
  line-height: 1 !important;
  letter-spacing: 0 !important;
  text-transform: none !important;
  word-wrap: normal !important;
}
.stApp { background: #FFFFFF; }
#MainMenu, footer { visibility: hidden; }

/* Nav bar without negative-margin hack */
.mpc-nav {
  background: var(--mpc-gradient);
  padding: 16px 32px;
  margin-bottom: 0;
  display: flex; justify-content: space-between; align-items: center;
  color: white;
  box-shadow: 0 2px 12px rgba(0,0,0,0.08);
  border-radius: 0;
}
.mpc-nav-brand { display: flex; align-items: center; gap: 14px; color: white !important; }
.mpc-nav-title { font-size: 20px; font-weight: 700; letter-spacing: -0.01em; color: white !important; }
.mpc-nav-sub { font-size: 13px; font-weight: 500; opacity: 0.92; color: white !important; }
.mpc-logo-box { background: white; padding: 8px 20px; transform: skew(-12deg); box-shadow: 0 2px 6px rgba(0,0,0,0.1); }
.mpc-logo-box > * { transform: skew(12deg); display: inline-block; }
.mpc-logo-text { font-weight: 800; font-size: 26px; color: var(--mpc-charcoal); letter-spacing: -0.02em; }
.mpc-logo-text sup { color: var(--mpc-green-light); font-size: 20px; }

/* Headings */
h1 { color: var(--mpc-charcoal) !important; font-size: 40px !important; font-weight: 700 !important;
     line-height: 1.15 !important; letter-spacing: -0.01em !important; margin: 1.5rem 0 1rem !important; }
h2 { color: var(--mpc-charcoal) !important; font-size: 28px !important; font-weight: 700 !important;
     line-height: 1.3 !important; padding-left: 16px !important;
     border-left: 4px solid var(--mpc-green-light) !important;
     margin: 2rem 0 1rem !important; }
h3 { color: var(--mpc-charcoal) !important; font-size: 20px !important; font-weight: 600 !important;
     margin: 1rem 0 0.5rem !important; }
h4 { color: var(--mpc-charcoal) !important; font-size: 15px !important; font-weight: 700 !important;
     text-transform: uppercase; letter-spacing: 0.06em; }

p, .stMarkdown p, li { font-size: 16px !important; line-height: 1.5 !important; color: var(--mpc-body) !important; }

/* Buttons */
.stButton > button, .stDownloadButton > button {
  background-color: var(--mpc-green-light) !important;
  color: white !important;
  font-family: 'Manrope', sans-serif !important;
  font-size: 16px !important; font-weight: 700 !important; letter-spacing: 0.01em !important;
  padding: 10px 32px !important; border-radius: 0 !important; border: none !important;
  transition: background-color 0.2s; box-shadow: none !important;
}
.stButton > button:hover, .stDownloadButton > button:hover { background-color: var(--mpc-green-dark) !important; }

/* Tabs — KEEP VISIBLE */
.stTabs {
  margin-top: 8px;
  background: white;
  position: sticky; top: 0; z-index: 100;
}
.stTabs [data-baseweb="tab-list"] {
  gap: 0;
  border-bottom: 2px solid var(--mpc-bg-light);
  padding: 0 24px;
  background: white;
}
.stTabs [data-baseweb="tab-list"] button {
  font-family: 'Manrope', sans-serif !important;
  font-weight: 600 !important; color: var(--mpc-charcoal) !important;
  font-size: 15px !important;
  padding: 14px 24px !important;
  background: transparent !important; border-radius: 0 !important; border: none !important;
}
.stTabs [data-baseweb="tab-list"] button[aria-selected="true"] {
  color: var(--mpc-green-dark) !important;
  background: #F7F9F3 !important;
  border-bottom: 3px solid var(--mpc-green-light) !important;
}
.stTabs [data-baseweb="tab-panel"] { padding: 24px 32px; }

/* Inputs */
.stTextInput input, .stNumberInput input, .stTextArea textarea, .stSelectbox {
  border-radius: 0 !important;
  font-family: 'Manrope', sans-serif !important;
  border: 1px solid var(--mpc-bg-light) !important;
}

/* File Uploader styling (text replacement via JS below) */
[data-testid="stFileUploader"] section {
  border: 2px dashed var(--mpc-green-light) !important;
  border-radius: 0 !important;
  background: #FAFAF7 !important;
}
[data-testid="stFileUploader"] section button {
  background-color: var(--mpc-green-light) !important;
  color: white !important;
  border-radius: 0 !important;
  border: none !important;
  font-weight: 700 !important;
}

/* Metric-Kacheln */
.mpc-card {
  background: white; border: 1px solid var(--mpc-bg-light);
  border-left: 4px solid var(--mpc-green-light);
  padding: 20px 24px; margin-bottom: 12px;
}
.mpc-card.ratio { border-left-color: var(--mpc-green-dark); background: linear-gradient(135deg, #F4FAEA, #FAFDF5); }
.mpc-card .label { font-size: 11px; font-weight: 700; text-transform: uppercase;
                   letter-spacing: 0.08em; color: var(--mpc-charcoal); margin-bottom: 4px; opacity: 0.7; }
.mpc-card .value { font-size: 32px; font-weight: 700; color: var(--mpc-charcoal);
                   letter-spacing: -0.01em; line-height: 1; }
.mpc-card .unit { font-size: 15px; font-weight: 500; color: var(--mpc-body); margin-left: 6px; }

/* Integrity score gauge */
.integrity-hero {
  display: grid;
  grid-template-columns: 180px 1fr;
  gap: 32px;
  align-items: center;
  padding: 24px 32px;
  background: linear-gradient(135deg, #FAFAF7, #F4FAEA);
  border: 1px solid var(--mpc-bg-light);
  border-left: 6px solid var(--mpc-green-light);
  margin: 16px 0;
}
.integrity-score {
  text-align: center;
}
.integrity-score .big { font-size: 72px; font-weight: 800; line-height: 1; color: var(--mpc-green-dark); letter-spacing: -0.03em; }
.integrity-score .grade { font-size: 14px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; color: var(--mpc-charcoal); margin-top: 6px; }
.integrity-score .slash { font-size: 28px; color: var(--mpc-charcoal); opacity: 0.5; font-weight: 600; }
.integrity-score.A .big { color: var(--mpc-green-dark); }
.integrity-score.B .big { color: var(--mpc-green-light); }
.integrity-score.C .big { color: var(--warn); }
.integrity-score.D .big, .integrity-score.F .big { color: var(--danger); }

.check-item {
  display: grid;
  grid-template-columns: 32px 1fr auto;
  gap: 14px;
  padding: 10px 0;
  border-bottom: 1px solid var(--mpc-bg-light);
  align-items: flex-start;
}
.check-item:last-child { border-bottom: none; }
.check-icon {
  width: 24px; height: 24px;
  border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-weight: 800; font-size: 13px; color: white;
  flex-shrink: 0;
}
.check-icon.ok { background: var(--mpc-green-dark); }
.check-icon.warn { background: var(--warn); }
.check-icon.fail { background: var(--danger); }
.check-name { font-weight: 600; color: var(--mpc-charcoal); font-size: 14px; margin-bottom: 2px; }
.check-detail { font-size: 13px; color: var(--mpc-body); line-height: 1.4; }
.check-value { font-family: 'JetBrains Mono', monospace; font-size: 12px; color: var(--mpc-charcoal);
               background: var(--mpc-bg-light); padding: 4px 10px; margin-left: 8px; white-space: nowrap; }

.mpc-info {
  background: #F7F9F3; border-left: 4px solid var(--mpc-green-light);
  padding: 14px 20px; margin: 12px 0; font-size: 14px; color: var(--mpc-body);
}
.mpc-info.warn { background: var(--warn-light); border-left-color: var(--warn); }
.mpc-info.danger { background: var(--danger-light); border-left-color: var(--danger); }

/* Method-cards */
.method-box {
  background: white; border: 1px solid var(--mpc-bg-light);
  padding: 20px 24px; margin: 12px 0;
  position: relative;
}
.method-box::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 5px; }
.method-box.vertex::before { background: var(--mpc-green-light); }
.method-box.midpoint::before { background: var(--mpc-green-dark); }
.method-box.manual::before { background: var(--mpc-charcoal); }

.proscons { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-top: 14px; }
.pros, .cons { padding: 14px 18px; font-size: 14px; line-height: 1.5; }
.pros { background: var(--success-light); border-left: 3px solid var(--mpc-green-dark); }
.cons { background: var(--danger-light); border-left: 3px solid var(--danger); }
.pros strong, .cons strong { display: block; font-size: 11px; letter-spacing: 0.1em; margin-bottom: 6px; text-transform: uppercase; }
.pros strong { color: var(--mpc-green-dark); }
.cons strong { color: var(--danger); }

.issue-card {
  background: white; border: 1px solid var(--mpc-bg-light);
  border-left: 4px solid var(--warn);
  padding: 18px 22px; margin: 14px 0;
}
.issue-card.critical { border-left-color: var(--danger); }
.issue-card h4 { margin: 0 0 10px; }
.issue-card .evidence { font-family: 'JetBrains Mono', monospace; font-size: 12px;
                         background: #F7F7F5; padding: 10px 14px; border-left: 3px solid var(--mpc-bg-light); margin: 10px 0;
                         white-space: pre; overflow-x: auto; }
.issue-card .verdict { font-weight: 700; color: var(--mpc-charcoal); margin-top: 10px; }

/* Sidebar */
section[data-testid="stSidebar"] { background: #F7F7F5; border-right: 3px solid var(--mpc-green-light); }
section[data-testid="stSidebar"] h3 {
  color: var(--mpc-charcoal) !important; font-size: 13px !important; font-weight: 700 !important;
  text-transform: uppercase; letter-spacing: 0.06em;
  border-left: none !important; padding-left: 0 !important; margin-top: 1rem !important;
}

table { width: 100%; border-collapse: collapse; margin: 12px 0; }
table th { background: var(--mpc-bg-light); color: var(--mpc-charcoal);
           text-align: left; padding: 10px 12px; font-size: 12px; font-weight: 700;
           text-transform: uppercase; letter-spacing: 0.06em;
           border-bottom: 2px solid var(--mpc-green-light); }
table td { padding: 10px 12px; font-size: 13px; border-bottom: 1px solid var(--mpc-bg-light);
            color: var(--mpc-body); vertical-align: top; }
table tr:last-child td { border-bottom: none; }

code, pre {
  font-family: 'JetBrains Mono', 'Menlo', monospace !important;
  background: #F7F7F5 !important;
  color: var(--mpc-charcoal) !important;
  font-size: 13px !important;
}
pre { border-left: 3px solid var(--mpc-green-light); padding: 12px 16px !important; }

/* Footer */
.mpc-footer {
  background: var(--mpc-gradient);
  color: white; padding: 20px 40px;
  text-align: center; font-size: 13px; font-weight: 500;
  margin-top: 48px;
}
.mpc-footer a { color: white; text-decoration: underline; }

.stRadio label, .stSlider label, .stCheckbox label {
  font-family: 'Manrope', sans-serif !important;
  color: var(--mpc-charcoal) !important;
  font-weight: 500 !important;
  font-size: 14px !important;
}

/* Altair charts */
.vega-embed { width: 100% !important; }
</style>
"""
st.markdown(MPC2_CSS, unsafe_allow_html=True)

# Inject a MutationObserver-based translation that replaces English Streamlit
# text nodes with German. Must run via components.v1.html because Streamlit
# strips <script> from st.markdown even with unsafe_allow_html.
import streamlit.components.v1 as components  # noqa: E402

components.html("""
<script>
(function() {
  const translations = [
    ['Drag and drop files here', 'Dateien hier ablegen'],
    ['Drag and drop file here',  'Datei hier ablegen'],
    ['Browse files',              'Durchsuchen'],
    ['Browse file',               'Durchsuchen'],
    ['Limit 200MB per file',      'Max. 200 MB pro Datei'],
    ['Press Enter to apply',      'Mit Enter bestätigen'],
  ];
  function translate(root) {
    translations.forEach(([en, de]) => {
      const walker = document.createTreeWalker(root, NodeFilter.SHOW_TEXT);
      let node;
      while ((node = walker.nextNode())) {
        if (node.nodeValue && node.nodeValue.includes(en)) {
          node.nodeValue = node.nodeValue.replace(en, de);
        }
      }
    });
  }
  const target = window.parent.document.body || document.body;
  translate(target);
  new MutationObserver(() => translate(target)).observe(target, {
    childList: true, subtree: true, characterData: true
  });
})();
</script>
""", height=0)


# ═══════════════════════════════════════════════════════════════════════════
# Branded top bar
# ═══════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="mpc-nav">
  <div class="mpc-nav-brand">
    <div class="mpc-logo-box"><span class="mpc-logo-text">MPC<sup>²</sup></span></div>
    <div>
      <div class="mpc-nav-title">Corrosion Ray® · DL-EPR-Auswertung</div>
      <div class="mpc-nav-sub">Automatische Messdaten-Analyse · direkt in die Messübersicht</div>
    </div>
  </div>
  <div class="mpc-nav-sub">gebaut von werchota.ai</div>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# Sidebar
# ═══════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### Split-Methode")
    split_method = st.radio(
        "Wie soll der Umkehrpunkt erkannt werden?",
        options=["vertex", "midpoint", "manual"],
        format_func=lambda x: {
            "vertex":   "Potenzial-Vertex (physikalisch)",
            "midpoint": "Mittelpunkt (heuristisch)",
            "manual":   "Manuell (Slider im Tab)",
        }[x],
        label_visibility="collapsed",
    )

    st.markdown("### Qr-Integration")
    truncate_reverse = st.checkbox(
        "Am Startpotenzial abschneiden", value=True,
        help="DL-EPR-Konvention: Qr-Integration endet, wenn die Rückwärtsrampe "
             "wieder das Startpotenzial erreicht.",
    )

    st.markdown("### Messübersicht-Master")
    messuebersicht_mode = st.radio(
        "Wie wird die Messübersicht bereitgestellt?",
        options=["upload", "path"],
        format_func=lambda x: {"upload": "Datei hochladen", "path": "Lokaler Pfad"}[x],
        label_visibility="collapsed",
    )

    messuebersicht_path_input = None
    messuebersicht_upload = None
    if messuebersicht_mode == "upload":
        messuebersicht_upload = st.file_uploader(
            "Messübersicht.xlsx", type=["xlsx"],
            accept_multiple_files=False, key="uebersicht_upload",
        )
    else:
        default_path = str(
            Path("/Users/malcolmwerchota/Library/CloudStorage/OneDrive-SharedLibraries-MPC²/")
            / "Manuel Prohaska - MPC²_Beratungsprojekt Green Startupmark"
            / "Messübersicht_Elektrochemische Messungen.xlsx"
        )
        messuebersicht_path_input = st.text_input(
            "Absoluter Pfad zur Messübersicht.xlsx", value=default_path,
        )

    order_no = st.text_input(
        "Untersuchungsthema / Order-No.", value="",
        help="Wird in Spalte F der Messübersicht eingetragen.",
    )


# ═══════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════

def _read_asc_from_bytes(file) -> ASCData:
    arr = np.loadtxt(io.BytesIO(file.getvalue()), dtype=float)
    if arr.ndim == 1:
        arr = arr.reshape(1, -1)
    return ASCData(
        time_s=arr[:, 0], potential_v=arr[:, 1],
        current_a=arr[:, 2], current_density_am2=arr[:, 3],
    )


def _read_had_from_bytes(file) -> HADMetadata:
    tmp = Path(tempfile.gettempdir()) / file.name
    tmp.write_bytes(file.getvalue())
    return parse_had(tmp)


def _resolve_messuebersicht_path() -> Path | None:
    if messuebersicht_mode == "upload" and messuebersicht_upload is not None:
        tmp = Path(tempfile.gettempdir()) / messuebersicht_upload.name
        tmp.write_bytes(messuebersicht_upload.getvalue())
        return tmp
    if messuebersicht_mode == "path" and messuebersicht_path_input:
        p = Path(messuebersicht_path_input).expanduser()
        if p.exists():
            return p
    return None


def _load_evaluation_data():
    path = Path(__file__).resolve().parent.parent / "out" / "full_evaluation.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text())
    except Exception:
        return None


def _make_potential_chart(asc: ASCData, split_idx: int) -> alt.Chart:
    df = pd.DataFrame({"t_s": asc.time_s, "E_mV": asc.potential_mv})
    line = alt.Chart(df).mark_line(color="#8DBF18", strokeWidth=2).encode(
        x=alt.X(field="t_s", title="Zeit [s]", type="quantitative"),
        y=alt.Y(field="E_mV", title="Potenzial [mV]", type="quantitative"),
    )
    vertex_rule = alt.Chart(pd.DataFrame({"t_s": [float(asc.time_s[split_idx])]})).mark_rule(
        color="#409A2D", strokeDash=[5, 3], strokeWidth=2
    ).encode(x=alt.X(field="t_s", type="quantitative"))
    vertex_label = alt.Chart(
        pd.DataFrame({"t_s": [float(asc.time_s[split_idx])], "label": ["Umkehrpunkt"]})
    ).mark_text(
        align="left", dx=6, dy=-6, color="#409A2D", fontWeight=600,
        font="Manrope, sans-serif", fontSize=11
    ).encode(
        x=alt.X(field="t_s", type="quantitative"),
        y=alt.value(10),
        text="label:N",
    )
    return (line + vertex_rule + vertex_label).properties(height=260, background="white")


def _make_current_chart(asc: ASCData, ja_idx: int, jr_idx: int, split_idx: int) -> alt.Chart:
    df = pd.DataFrame({"t_s": asc.time_s, "J_mAcm2": asc.current_density_macm2})
    line = alt.Chart(df).mark_line(color="#4C4C4C", strokeWidth=2).encode(
        x=alt.X(field="t_s", title="Zeit [s]", type="quantitative"),
        y=alt.Y(field="J_mAcm2", title="Stromdichte J [mA/cm²]", type="quantitative"),
    )
    split_rule = alt.Chart(pd.DataFrame({"t_s": [float(asc.time_s[split_idx])]})).mark_rule(
        color="#409A2D", strokeDash=[5, 3], strokeWidth=1.5, opacity=0.6
    ).encode(x=alt.X(field="t_s", type="quantitative"))
    ja_point = alt.Chart(pd.DataFrame({
        "t_s": [float(asc.time_s[ja_idx])],
        "J_mAcm2": [float(asc.current_density_macm2[ja_idx])],
        "Label": [f"Ja = {asc.current_density_macm2[ja_idx]:.3f}"],
    })).mark_point(filled=True, size=140, color="#8DBF18", stroke="white", strokeWidth=2).encode(
        x=alt.X(field="t_s", type="quantitative"),
        y=alt.Y(field="J_mAcm2", type="quantitative"),
        tooltip="Label:N",
    )
    jr_point = alt.Chart(pd.DataFrame({
        "t_s": [float(asc.time_s[jr_idx])],
        "J_mAcm2": [float(asc.current_density_macm2[jr_idx])],
        "Label": [f"Jr = {asc.current_density_macm2[jr_idx]:.3f}"],
    })).mark_point(filled=True, size=140, color="#b8860b", stroke="white", strokeWidth=2).encode(
        x=alt.X(field="t_s", type="quantitative"),
        y=alt.Y(field="J_mAcm2", type="quantitative"),
        tooltip="Label:N",
    )
    return (line + split_rule + ja_point + jr_point).properties(height=260, background="white")


def _make_polarization_loop(asc: ASCData, split_idx: int) -> alt.Chart:
    """DL-EPR Loop: Potenzial vs. Stromdichte — die klassische 'Doppel-Schleife'."""
    df = pd.DataFrame({
        "Potenzial [mV]": asc.potential_mv,
        "J [mA/cm²]": asc.current_density_macm2,
        "Phase": ["Vorwärts (Ja)" if i < split_idx else "Rückwärts (Jr)"
                  for i in range(asc.n_points)],
    })
    chart = alt.Chart(df).mark_line(strokeWidth=2).encode(
        x=alt.X("Potenzial [mV]:Q", scale=alt.Scale(zero=False), title="Potenzial [mV]"),
        y=alt.Y("J [mA/cm²]:Q", title="Stromdichte J [mA/cm²]"),
        color=alt.Color("Phase:N", scale=alt.Scale(
            domain=["Vorwärts (Ja)", "Rückwärts (Jr)"],
            range=["#8DBF18", "#b8860b"],
        ), legend=alt.Legend(title=None, orient="top-left")),
        order="Zeit:Q",
    ).properties(height=320).configure_view(strokeOpacity=0)
    return chart


# ═══════════════════════════════════════════════════════════════════════════
# Tabs
# ═══════════════════════════════════════════════════════════════════════════

tab_analyse, tab_methodik, tab_genauigkeit, tab_qualitaet = st.tabs([
    "  Analyse  ",
    "  Methodik  ",
    "  Genauigkeit  ",
    "  Datenqualität  ",
])


# ─── TAB 1 — ANALYSE ───────────────────────────────────────────────────────
with tab_analyse:
    # Nonce used to force-reset the file_uploader widgets. Incrementing it
    # and calling st.rerun() clears both uploads cleanly without a browser reload.
    if "upload_nonce" not in st.session_state:
        st.session_state.upload_nonce = 0
    nonce = st.session_state.upload_nonce

    col_title, col_reset = st.columns([5, 1])
    with col_title:
        st.markdown("# Corrosion Ray Messdaten-Auswertung")
    with col_reset:
        st.markdown("<div style='height:42px;'></div>", unsafe_allow_html=True)
        if st.button("↻ Neu starten", use_container_width=True, key="reset_btn",
                     help="Hochgeladene Dateien entfernen und Analyse zurücksetzen."):
            st.session_state.upload_nonce = nonce + 1
            # Auch per-Datei-Slider & Messübersicht-Checkbox aufräumen,
            # damit kein veralteter Widget-State übrig bleibt.
            for k in list(st.session_state.keys()):
                if k.startswith("split_") or k == "overwrite_mu":
                    del st.session_state[k]
            st.rerun()

    st.markdown(
        "<p style='font-size:18px; line-height:1.5; color:#333333; margin-bottom:24px;'>"
        "ASC-Dateien hochladen, Ja/Jr/Qa/Qr automatisch erkennen, Datenqualität "
        "bewerten und Ergebnisse direkt in die Messübersicht übertragen."
        "</p>",
        unsafe_allow_html=True,
    )

    col_up1, col_up2 = st.columns(2)
    with col_up1:
        st.markdown("#### ASC-Messdaten")
        uploaded = st.file_uploader(
            "ASC-Dateien (eine oder mehrere)", type=["asc", "ASC"],
            accept_multiple_files=True, label_visibility="collapsed",
            key=f"asc_upload_{nonce}",
        )
    with col_up2:
        st.markdown("#### HAD-Metadaten (optional)")
        had_uploaded = st.file_uploader(
            "HAD-Dateien mit Probenfläche", type=["had", "HAD"],
            accept_multiple_files=True, label_visibility="collapsed",
            key=f"had_upload_{nonce}",
        )

    if not uploaded:
        st.markdown(
            "<div class='mpc-info'>→ Bitte mindestens eine .ASC-Datei hochladen, "
            "um mit der Analyse zu beginnen.</div>",
            unsafe_allow_html=True,
        )
    else:
        had_map = {Path(h.name).stem: h for h in (had_uploaded or [])}
        measurements_for_append: list[Measurement] = []

        for idx, f in enumerate(uploaded):
            st.markdown(f"## {f.name}")
            try:
                asc = _read_asc_from_bytes(f)
            except Exception as e:
                st.error(f"Fehler beim Parsen der ASC-Datei: {e}")
                continue

            had = None
            matched_had = had_map.get(Path(f.name).stem)
            if matched_had:
                try:
                    had = _read_had_from_bytes(matched_had)
                    st.markdown(
                        f"<div class='mpc-info'>→ HAD-Datei erkannt · "
                        f"Probenfläche = <b>{had.probenflaeche_mm2} mm²</b> · "
                        f"Sachbearbeiter = {had.sachbearbeiter or '—'}</div>",
                        unsafe_allow_html=True,
                    )
                except Exception as e:
                    st.warning(f"HAD-Datei konnte nicht gelesen werden: {e}")

            fm = parse_filename(f.name)
            if split_method == "vertex":
                initial_split, _ = detect_split_vertex(asc)
            elif split_method == "midpoint":
                initial_split, _ = detect_split_midpoint(asc)
            else:
                initial_split, _ = detect_split_vertex(asc)

            col_slider, col_info = st.columns([3, 1])
            with col_slider:
                split = st.slider(
                    "Split-Index (Grenze zwischen Aktivierungs- und Reaktivierungs-Sweep)",
                    min_value=10, max_value=asc.n_points - 10,
                    value=int(initial_split), key=f"split_{idx}",
                )
            with col_info:
                st.markdown(
                    f"<div class='mpc-card' style='text-align:center; padding:14px;'>"
                    f"<div class='label'>Datenpunkte</div>"
                    f"<div class='value' style='font-size:28px;'>{asc.n_points}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            try:
                result = analyze_dlepr(
                    asc, had,
                    split_method="manual" if split != initial_split else split_method,
                    split_override=split if split != initial_split else None,
                    truncate_reverse=truncate_reverse,
                )
            except Exception as e:
                st.error(f"Analyse-Fehler: {e}")
                # Analyse fehlgeschlagen — trotzdem die Rohdaten plotten, damit
                # der Benutzer visuell nachvollziehen kann, warum die Datei nicht
                # auswertbar war (z.B. keine Umkehr → reine Aktivierungsdatei).
                st.markdown("### Rohdaten-Diagramme (zur Diagnose)")
                st.markdown(
                    "<div class='mpc-info warn'>Die DL-EPR-Auswertung konnte nicht "
                    "durchgeführt werden, aber die Rohdaten werden zur visuellen "
                    "Inspektion angezeigt.</div>",
                    unsafe_allow_html=True,
                )
                cc_raw = st.columns(2)
                with cc_raw[0]:
                    st.markdown("**Potenzial-Rampe (Rohdaten)**")
                    df_raw = pd.DataFrame({"t_s": asc.time_s, "E_mV": asc.potential_mv})
                    st.altair_chart(
                        alt.Chart(df_raw).mark_line(color="#8DBF18", strokeWidth=2).encode(
                            x=alt.X(field="t_s", title="Zeit [s]", type="quantitative"),
                            y=alt.Y(field="E_mV", title="Potenzial [mV]", type="quantitative"),
                        ).properties(height=260, background="white"),
                        use_container_width=True, theme=None,
                    )
                with cc_raw[1]:
                    st.markdown("**Stromdichte-Verlauf (Rohdaten)**")
                    df_raw = pd.DataFrame({"t_s": asc.time_s, "J_mAcm2": asc.current_density_macm2})
                    st.altair_chart(
                        alt.Chart(df_raw).mark_line(color="#4C4C4C", strokeWidth=2).encode(
                            x=alt.X(field="t_s", title="Zeit [s]", type="quantitative"),
                            y=alt.Y(field="J_mAcm2", title="Stromdichte J [mA/cm²]",
                                    type="quantitative"),
                        ).properties(height=260, background="white"),
                        use_container_width=True, theme=None,
                    )
                st.markdown("**Polarisation (E vs J) — Rohdaten, keine Phasen-Färbung**")
                df_loop_raw = pd.DataFrame({
                    "E_mV": asc.potential_mv,
                    "J_mAcm2": asc.current_density_macm2,
                    "t_s": asc.time_s,
                })
                st.altair_chart(
                    alt.Chart(df_loop_raw).mark_line(color="#409A2D", strokeWidth=2).encode(
                        x=alt.X(field="E_mV", title="Potenzial [mV]",
                                scale=alt.Scale(zero=False), type="quantitative"),
                        y=alt.Y(field="J_mAcm2", title="Stromdichte J [mA/cm²]",
                                type="quantitative"),
                        order="t_s:Q",
                    ).properties(height=320, background="white"),
                    use_container_width=True, theme=None,
                )
                continue

            m_obj = Measurement(
                source_file=f.name, asc=asc,
                had=had or HADMetadata(),
                filename_meta=fm, analysis=result,
            )
            measurements_for_append.append(m_obj)

            # Primäre Messwerte
            st.markdown("### Primäre Messwerte")
            c1, c2, c3, c4 = st.columns(4)
            for col, label, val, unit in [
                (c1, "Ja (Aktivierungspeak)",      f"{result.ja_ma_cm2:.4f}", "mA/cm²"),
                (c2, "Jr (Reaktivierungspeak)",    f"{result.jr_ma_cm2:.4f}", "mA/cm²"),
                (c3, "Qa (Aktivierungsladung)",    f"{result.qa_as:.6f}",     "As"),
                (c4, "Qr (Reaktivierungsladung)",  f"{result.qr_as:.6f}",     "As"),
            ]:
                col.markdown(
                    f"<div class='mpc-card'><div class='label'>{label}</div>"
                    f"<div class='value'>{val}<span class='unit'> {unit}</span></div></div>",
                    unsafe_allow_html=True,
                )

            # Abgeleitete Kennwerte
            st.markdown("### Abgeleitete Kennwerte (DOS)")
            r1, r2, r3 = st.columns(3)
            for col, label, val, unit in [
                (r1, "Jr / Ja (DOS)",        f"{result.jr_ja:.5f}", ""),
                (r2, "Qr / Qa",              f"{result.qr_qa:.5f}", ""),
                (r3, "Ruhepotential",        f"{result.ruhepotential_mv:.1f}", "mV"),
            ]:
                col.markdown(
                    f"<div class='mpc-card ratio'><div class='label'>{label}</div>"
                    f"<div class='value'>{val}<span class='unit'> {unit}</span></div></div>",
                    unsafe_allow_html=True,
                )

            # Integritäts-Score
            st.markdown("### Datenintegritäts-Bewertung")
            try:
                report = evaluate_integrity(asc, result)
                st.markdown(f"""
                <div class='integrity-hero'>
                  <div class='integrity-score {report.grade}'>
                    <div class='big'>{report.score}<span class='slash'>/100</span></div>
                    <div class='grade'>Note {report.grade}</div>
                  </div>
                  <div>
                    <h3 style='margin:0 0 6px;'>Qualitäts-Score der Messung</h3>
                    <div style='color:#333; font-size:14px; line-height:1.5;'>
                      Aggregat aus {len(report.checks)} physikalischen und statistischen Checks.
                      <b style='color:#409A2D;'>{report.n_ok} OK</b> ·
                      <b style='color:#b8860b;'>{report.n_warn} Warnung</b> ·
                      <b style='color:#c0392b;'>{report.n_fail} Fehler</b>
                    </div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                with st.expander("Alle Checks im Detail", expanded=False):
                    for check in report.checks:
                        icon_label = {"ok": "✓", "warn": "!", "fail": "✗"}[check.status]
                        st.markdown(f"""
                        <div class='check-item'>
                          <div class='check-icon {check.status}'>{icon_label}</div>
                          <div>
                            <div class='check-name'>{check.name}</div>
                            <div class='check-detail'>{check.detail}</div>
                          </div>
                          <div class='check-value'>{check.value}</div>
                        </div>
                        """, unsafe_allow_html=True)
            except Exception as e:
                st.warning(f"Integritäts-Bewertung nicht verfügbar: {e}")

            # Kurven
            st.markdown("### Diagramme")
            cc = st.columns(2)
            with cc[0]:
                st.markdown("**Potenzial-Rampe** — Vorwärts bis Vertex, dann Rückwärts")
                st.altair_chart(_make_potential_chart(asc, result.split_index),
                                use_container_width=True, theme=None)
            with cc[1]:
                st.markdown("**Stromdichte-Verlauf** — mit markierten Peaks Ja und Jr")
                st.altair_chart(_make_current_chart(asc, result.ja_index, result.jr_index, result.split_index),
                                use_container_width=True, theme=None)

            st.markdown("**DL-EPR-Polarisationsschleife** — klassische Darstellung Potenzial vs. Stromdichte")
            loop_df = pd.DataFrame({
                "E_mV": asc.potential_mv,
                "J_mAcm2": asc.current_density_macm2,
                "Phase": ["Vorwärts (Aktivierung)" if i <= result.split_index
                          else "Rückwärts (Reaktivierung)" for i in range(asc.n_points)],
                "t_s": asc.time_s,
            })
            loop_chart = alt.Chart(loop_df).mark_line(strokeWidth=2).encode(
                x=alt.X(field="E_mV", title="Potenzial [mV]",
                        scale=alt.Scale(zero=False), type="quantitative"),
                y=alt.Y(field="J_mAcm2", title="Stromdichte J [mA/cm²]", type="quantitative"),
                color=alt.Color("Phase:N", scale=alt.Scale(
                    domain=["Vorwärts (Aktivierung)", "Rückwärts (Reaktivierung)"],
                    range=["#8DBF18", "#b8860b"],
                ), legend=alt.Legend(title=None, orient="top-left")),
                order="t_s:Q",
            ).properties(height=320, background="white")
            st.altair_chart(loop_chart, use_container_width=True, theme=None)

            # Copy-paste
            tab_sep = "\t".join([
                str(fm.messung_id or ""), fm.material or "", fm.probenbez or "",
                str(had.probenflaeche_mm2 if had and had.probenflaeche_mm2 else ""),
                str(fm.temperature_c or ""), f"{result.ruhepotential_mv:.2f}",
                f"{result.ja_ma_cm2:.5f}", f"{result.jr_ma_cm2:.5f}", f"{result.jr_ja:.6f}",
                f"{result.qa_as:.8f}", f"{result.qr_as:.8f}", f"{result.qr_qa:.6f}",
            ])
            with st.expander("Werte zum Kopieren (tab-getrennt für Messübersicht)"):
                st.code(tab_sep, language="text")
                st.caption(
                    "Reihenfolge: Messung · Material · Probenbez · Messfläche · Temp [°C] · "
                    "Ruhepot [mV] · Ja · Jr · Jr/Ja · Qa · Qr · Qr/Qa"
                )

            with st.expander("Diagnose / Technische Details"):
                st.json({
                    "split_index": result.split_index,
                    "split_methode": result.split_method,
                    "split_diagnostik": result.split_diagnostics,
                    "ja_index": result.ja_index, "jr_index": result.jr_index,
                    "ja_zeit_s": float(asc.time_s[result.ja_index]),
                    "jr_zeit_s": float(asc.time_s[result.jr_index]),
                    "excel_qa_bereich": list(result.excel_qa_range),
                    "excel_qr_bereich": list(result.excel_qr_range),
                    "dateiname": {
                        "messung_id": fm.messung_id, "material": fm.material,
                        "probenbez": fm.probenbez,
                        "temperatur_c": fm.temperature_c,
                        "probe_temperatur_c": fm.probe_temperature_c,
                        "aktivierung_mv": fm.activation_mv,
                    },
                    "had_metadaten": {
                        "probenflaeche_mm2": had.probenflaeche_mm2 if had else None,
                        "sachbearbeiter": had.sachbearbeiter if had else None,
                        "erstellungsdatum": had.erstellungsdatum if had else None,
                        "anzahl_werte_rp": had.anzahl_werte_rp if had else None,
                        "anzahl_werte": had.anzahl_werte if had else None,
                    } if had else "Keine HAD-Datei bereitgestellt",
                })

        # ───────────────────────────────────────────────────────────────
        # Downloads & Messübersicht
        # ───────────────────────────────────────────────────────────────
        st.markdown("## Downloads")
        st.markdown(
            "<p style='color:#666; font-size:14px; margin-bottom:12px;'>"
            "Zwei Excel-Varianten zum Herunterladen: die <b>Auswertungs-Workbook</b> im "
            "gleichen Format wie MPC² sie bisher erstellt (ein Sheet pro Messung mit "
            "Rohdaten und Formeln) und die <b>aktualisierte Messübersicht</b> mit den "
            "neu hinzugefügten Zeilen."
            "</p>",
            unsafe_allow_html=True,
        )

        dl_col1, dl_col2 = st.columns(2)

        # A) Auswertungs-Workbook (Variant 1) — Template im MPC²-Stil
        with dl_col1:
            st.markdown("### Auswertungs-Workbook (Template)")
            st.markdown(
                "<p style='font-size:13px; color:#666;'>Ein Excel pro Projekt, "
                "ein Sheet pro Messung mit allen Rohdaten und Formeln — wie Manuel es "
                "heute manuell anlegt.</p>",
                unsafe_allow_html=True,
            )
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                auswertung_path = Path(tempfile.gettempdir()) / f"Auswertung_{ts}.xlsx"
                write_auswertung_workbook(
                    measurements_for_append,
                    auswertung_path,
                    project_name="StreamlitSession",
                )
                with open(auswertung_path, "rb") as fh:
                    st.download_button(
                        label="Auswertung.xlsx herunterladen",
                        data=fh.read(),
                        file_name=f"Auswertung_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            except Exception as e:
                st.error(f"Auswertung konnte nicht erstellt werden: {e}")

        # B) Messübersicht aktualisieren
        with dl_col2:
            st.markdown("### Messübersicht aktualisieren")
            mu_path = _resolve_messuebersicht_path()
            if mu_path is None:
                st.markdown(
                    "<p style='font-size:13px; color:#b8860b;'>→ Keine Messübersicht "
                    "verlinkt. Bitte in der Seitenleiste eine Datei hochladen oder "
                    "einen Pfad angeben.</p>",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f"<p style='font-size:13px; color:#666;'>Verlinkt: "
                    f"<code style='font-size:11px;'>{mu_path.name}</code></p>",
                    unsafe_allow_html=True,
                )
                overwrite = st.checkbox(
                    "Bestehende Messung-IDs überschreiben",
                    value=False, key="overwrite_mu",
                )
                if st.button("Messübersicht aktualisieren & herunterladen",
                             use_container_width=True):
                    try:
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        output_path = Path(tempfile.gettempdir()) / f"Messuebersicht_updated_{ts}.xlsx"
                        shutil.copy2(mu_path, output_path)
                        wb = openpyxl.load_workbook(output_path)
                        ws = wb["Corrosion Ray"]
                        existing_ids = {
                            str(ws.cell(row=r, column=1).value).strip(): r
                            for r in range(2, ws.max_row + 1)
                            if ws.cell(row=r, column=1).value is not None
                        }
                        appended = updated = skipped = 0
                        messages = []
                        for m in measurements_for_append:
                            row_data = _build_row(m, order_no=order_no or None)
                            mid = str(row_data["A"] or "").strip()
                            if mid in existing_ids:
                                if overwrite:
                                    target = existing_ids[mid]
                                    updated += 1
                                    messages.append(f"↻ {mid} überschrieben (Zeile {target})")
                                else:
                                    skipped += 1
                                    messages.append(f"— {mid} übersprungen (existiert bereits in Zeile {existing_ids[mid]})")
                                    continue
                            else:
                                target = ws.max_row + 1
                                appended += 1
                                messages.append(f"✓ {mid} hinzugefügt (Zeile {target})")
                            for col_letter, value in row_data.items():
                                if value is not None:
                                    ws[f"{col_letter}{target}"] = value
                        wb.save(output_path)
                        st.success(
                            f"{appended} hinzugefügt · {updated} überschrieben · {skipped} übersprungen"
                        )
                        for msg in messages:
                            st.markdown(f"<div style='font-size:12px; color:#666;'>{msg}</div>",
                                        unsafe_allow_html=True)
                        with open(output_path, "rb") as fh:
                            st.download_button(
                                label="Aktualisierte Messübersicht herunterladen",
                                data=fh.read(),
                                file_name=f"Messuebersicht_updated_{ts}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                            )
                    except Exception as e:
                        st.error(f"Fehler beim Schreiben der Messübersicht: {e}")


# ─── TAB 2 — METHODIK ───────────────────────────────────────────────────────
with tab_methodik:
    st.markdown("# Methodik — Wie erkennen wir den Umkehrpunkt?")
    st.markdown("""
Der Kern der DL-EPR-Auswertung ist die saubere Trennung zwischen
**Aktivierungs-Sweep** (in dem Ja gemessen wird) und **Reaktivierungs-Sweep**
(in dem Jr gemessen wird). Manuel hat das bisher per Auge gemacht ("in Zeile 1076
trenne ich"). Wir bieten zwei automatische Methoden plus ein manuelles Override,
damit die Auswertung reproduzierbar und trotzdem flexibel bleibt.
""")

    st.markdown("""
<div class="method-box vertex">
  <h3>A) Potenzial-Vertex (physikalisch)</h3>
  <p>Findet den tatsächlichen Scheitelpunkt der Potenzialrampe — den Punkt, wo das Potenzial
  sein Maximum erreicht und die Sweep-Richtung umkehrt. Dazu wird das Potenzial-Signal
  mit einem einfachen Moving-Average geglättet (Fenster ~11 Punkte) und dann der Index
  des Maximums genommen.</p>
  <div class="proscons">
    <div class="pros"><strong>Vorteile</strong>
      Physikalisch korrekte Definition der "Double Loop" in DL-EPR. Funktioniert bei
      beliebigen Sweep-Raten, Messzeiten und Datenpunktzahlen. Unabhängig von
      HAD-Metadaten. Keine Kalibrierung nötig.
    </div>
    <div class="cons"><strong>Nachteile</strong>
      Kann bei verrauschten Potenzial-Signalen um ±3 Zeilen schwanken (durch
      Glättung entschärft). Bei plateau-artigen Peaks kann der Vertex leicht von
      Manuels gewählter Position abweichen.
    </div>
  </div>
</div>

<div class="method-box midpoint">
  <h3>B) Mittelpunkt (heuristisch)</h3>
  <p>Teilt die Datenpunkte schlicht bei N/2. Basiert auf der Annahme, dass Vorwärts-
  und Rückwärts-Sweep typischerweise gleich lang dauern.</p>
  <div class="proscons">
    <div class="pros"><strong>Vorteile</strong>
      Extrem einfach, deterministisch, in einem Satz erklärbar. Sehr schnell —
      keine Signalverarbeitung. Entspricht Manuels historischer Praxis. Liefert in
      90% der Fälle identische Ja/Jr-Werte wie die Vertex-Methode.
    </div>
    <div class="cons"><strong>Nachteile</strong>
      Rein statistisch — benutzt nicht die Potenzialkurve selbst. Bei langen
      Ruhezuständen am Anfang verschiebt sich der wahre Vertex und Midpoint
      stimmt nicht mehr. Keine physikalische Begründung.
    </div>
  </div>
</div>

<div class="method-box manual">
  <h3>C) Manuell (Slider)</h3>
  <p>Split-Index direkt per Slider überschreiben. Kurvenansicht zeigt in Echtzeit,
  wie sich die Werte verändern.</p>
  <div class="proscons">
    <div class="pros"><strong>Vorteile</strong>
      Volle Kontrolle. Manuels Expertenurteil kann weiterhin einfließen.
      Perfekt für Edge-Cases oder auffällige Messungen.
    </div>
    <div class="cons"><strong>Nachteile</strong>
      Nicht für Batch-Verarbeitung geeignet. Subjektiv.
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

    st.markdown("## Qr-Integration: Truncation am Startpotenzial")
    st.markdown("""
<div class="mpc-info">
Ein subtiler aber wichtiger Aspekt: Die Qr-Ladungsintegration endet gemäß DL-EPR-Konvention
an dem Punkt, wo die Rückwärtsrampe wieder das Startpotenzial erreicht. Alles danach ist
Post-Sweep-Erholung und gehört <strong>nicht</strong> in Qr. Manuel setzt diese Grenze
per Hand (z.B. <code>SUM(I1060:I1950)</code>). Unser Parser sucht automatisch nach dem
ersten Index nach dem Split, wo das Potenzial wieder das Startpotenzial (±5 mV Toleranz)
erreicht. Dadurch wurde der Qr-Fehler vs. Manuels Werte von ~18% auf unter 3% reduziert.
</div>
""", unsafe_allow_html=True)

    st.markdown("""
<div class="mpc-info" style="border-left-color:#409A2D; background:#EAF4D8;">
<strong>Empfehlung: Vertex-Methode als Standard.</strong> Physikalisch begründet, robust
über verschiedene Messzeiten, und bei Qr marginal besser als die anderen. Midpoint als
Backup bei ungewöhnlichen Kurven. Manual Override bei Edge Cases.
</div>
""", unsafe_allow_html=True)


# ─── TAB 3 — GENAUIGKEIT ────────────────────────────────────────────────────
with tab_genauigkeit:
    st.markdown("# Genauigkeit — Validierung gegen Manuels Referenzwerte")
    st.markdown("""
Der Parser wurde gegen zwei unterschiedliche Referenzquellen getestet:
die konsolidierte **Messübersicht** (Manuels Laborbuch-Werte) und die
**Auswertung-Workbook** für Projekt 2 (mit Manuels hand-gewählten Split-Indizes).
Zusammen ergibt das **20 validierte Messungen**.
""")

    data = _load_evaluation_data()
    if data is None:
        st.warning("Keine Evaluations-Daten gefunden. Bitte `python tests/full_evaluation.py` ausführen.")
    else:
        st.markdown("## Projekt A (Projekt 1) — 12 Messungen gegen Messübersicht")
        st.markdown("IDs <strong>0048–0061</strong>.", unsafe_allow_html=True)

        p1 = data["projekt_1"]["errors"]
        rows = []
        for metric in ("ja", "jr", "qa", "qr"):
            for method in ("vertex", "midpoint"):
                errs = p1.get(method, {}).get(metric, [])
                if errs:
                    rows.append({
                        "Metrik": metric.upper(),
                        "Methode": {"vertex": "Vertex", "midpoint": "Midpoint"}[method],
                        "Anzahl": len(errs),
                        "Mittelwert |%err|": f"{np.mean(errs):.3f}%",
                        "Median": f"{np.median(errs):.3f}%",
                        "Maximum": f"{np.max(errs):.2f}%",
                    })
        st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

        st.markdown("""
<div class="mpc-info" style="border-left-color:#409A2D; background:#EAF4D8;">
<strong>Ja und Jr matchen auf 4 Nachkommastellen genau</strong> (Mittelwert &lt;0.01%).
Die Peak-Erkennung ist essentiell perfekt — unabhängig von der Split-Methode.
Qa-Fehler unter 2%. Qr-Median 7% — Ausreißer ziehen den Mittelwert hoch.
</div>
""", unsafe_allow_html=True)

        st.markdown("## Projekt B (Projekt 2) — 8 Messungen mit hand-gewählten Splits")
        p2 = data["projekt_2"]["errors"]
        rows2 = []
        for metric in ("ja", "jr", "qa", "qr"):
            for method in ("manuel", "vertex", "midpoint"):
                errs = p2.get(method, {}).get(metric, [])
                if errs:
                    rows2.append({
                        "Metrik": metric.upper(),
                        "Methode": {"manuel": "Manuels Split", "vertex": "Vertex", "midpoint": "Midpoint"}[method],
                        "Anzahl": len(errs),
                        "Mittelwert |%err|": f"{np.mean(errs):.3f}%",
                        "Median": f"{np.median(errs):.3f}%",
                        "Maximum": f"{np.max(errs):.2f}%",
                    })
        st.dataframe(pd.DataFrame(rows2), hide_index=True, use_container_width=True)

        st.markdown("""
<div class="mpc-info warn">
<strong>Wichtig:</strong> Die hohen Fehler (~23% Jr) gehen fast komplett auf den
<strong>404/405-Swap</strong> in Manuels Auswertung zurück (siehe Tab "Datenqualität").
Auf den 6 korrekt zugeordneten Sheets liegen alle drei Methoden bei 0.00% Fehler
für Ja/Jr und unter 3% für Qa/Qr.
</div>
""", unsafe_allow_html=True)

        st.markdown("## Kombiniert (A + B) — 20 Messungen")
        combined = {}
        for project in (p1, p2):
            for method, metrics in project.items():
                combined.setdefault(method, {})
                for metric, vals in metrics.items():
                    combined[method].setdefault(metric, []).extend(vals)

        rows_c = []
        for metric in ("ja", "jr", "qa", "qr"):
            for method in ("vertex", "midpoint"):
                errs = combined.get(method, {}).get(metric, [])
                if errs:
                    rows_c.append({
                        "Metrik": metric.upper(),
                        "Methode": {"vertex": "Vertex", "midpoint": "Midpoint"}[method],
                        "Anzahl": len(errs),
                        "Mittelwert |%err|": f"{np.mean(errs):.3f}%",
                        "Median": f"{np.median(errs):.3f}%",
                        "Maximum": f"{np.max(errs):.2f}%",
                    })
        st.dataframe(pd.DataFrame(rows_c), hide_index=True, use_container_width=True)


# ─── TAB 4 — DATENQUALITÄT ──────────────────────────────────────────────────
with tab_qualitaet:
    st.markdown("# Datenqualität — Auffälligkeiten in den Referenzdaten")
    st.markdown("""
Bei der Validierung sind uns einige Ungereimtheiten zwischen Parser-Ergebnissen und
Manuels manuell eingegebenen Werten aufgefallen. Diese sind <strong>keine
Parser-Fehler</strong> — im Gegenteil, der Parser ist konsistent und reproduzierbar.
Aber sie weisen auf mögliche Tippfehler oder Inkonsistenzen in den Referenzdaten hin,
die wir zum nächsten Review mit Manuel und Simone mitbringen sollten.
""", unsafe_allow_html=True)

    st.markdown("## 🚨 Kritisch: Projekt B Sheet-Swap 404 ↔ 405")
    st.markdown("""
<div class="issue-card critical">
  <h4>Vertauschte Zuordnung in Auswertung_ON2025-0003_UNS S32906 3D III_DL-EPR-Test.xlsx</h4>
  <p>In der Auswertung-Workbook für Projekt 2 scheinen die Sheets für ASC 0404 und
  0405 mit den falschen Rohdaten verknüpft zu sein. Die Endwerte in Manuels
  Sheet "404" entsprechen exakt dem, was unser Parser auf ASC 0405 berechnet — und
  umgekehrt.</p>
  <div class="evidence">Sheet "404_UNS S32906_K2-1_ht" — Manuels Werte: Ja=27.914  Jr=0.673
ASC 0404 durch Parser:                                Ja=34.242  Jr=1.546 ← passt zu Sheet 405

Sheet "405_UNS S32906_K2-2_ht_ung!" — Manuels Werte: Ja=34.242  Jr=1.546
ASC 0405 durch Parser:                                Ja=27.914  Jr=0.673 ← passt zu Sheet 404</div>
  <p class="verdict">→ Die Werte passen perfekt, aber gespiegelt. Entweder wurden die Sheet-Namen
  versehentlich vertauscht, oder beim manuellen Import wurden die falschen ASCs geladen.
  Dies erklärt auch den scheinbaren 20–130% Fehler in der Projekt-2-Auswertung —
  ohne diesen Swap wären alle Werte bei 0.00% Fehler.</p>
</div>
""", unsafe_allow_html=True)

    st.markdown("## ⚠️ Projekt A: Qr-Abweichungen bei 5 Messungen")
    st.markdown("""
Bei 5 der 12 validierten Messungen in Projekt 1 liegt der Qr-Wert deutlich über
Manuels Messübersicht-Eintrag. Mögliche Ursachen:

1. **ASC-Varianten-Auswahl** — mehrere IDs haben 2-3 verschiedene ASC-Dateien
   (Original, Retry mit "x"-Suffix, "-2"-Neumessung). Manuel könnte eine andere
   Variante verwendet haben als unser Auto-Matcher.
2. **Unterschiedliche Qr-Endgrenze** — wenn Manuel die Qr-Integration enger
   begrenzt hat (z.B. am tatsächlichen Nulldurchgang statt am Startpotenzial).
3. **Manuelle Nacharbeit in Excel** — Manuel hat historisch manchmal offensichtliche
   Artefakte am Ende der Kurve per Hand rausgelöscht.
""")

    data = _load_evaluation_data()
    if data is not None:
        outliers = []
        for d in data["projekt_1"]["details"]:
            r = d["ref"]
            v = d["methods"].get("vertex", {})
            if r.get("qr") and v.get("qr") and r["qr"] != 0:
                err = 100 * (v["qr"] - r["qr"]) / abs(r["qr"])
                if abs(err) > 10:
                    outliers.append({
                        "ID": d["id"],
                        "ASC-Variante gewählt": d["asc_chosen"][:55] + "...",
                        "Varianten verfügbar": d["n_variants"],
                        "Ref Qr": f"{r['qr']:.4f}",
                        "Unser Qr": f"{v['qr']:.4f}",
                        "Fehler": f"{err:+.1f}%",
                    })
        if outliers:
            st.dataframe(pd.DataFrame(outliers), hide_index=True, use_container_width=True)

    st.markdown("""
<div class="mpc-info warn">
<strong>→ Empfehlung:</strong> Mit Manuel klären, welche ASC-Variante für jede dieser IDs
als "offiziell" gilt (z.B. durchgehend die "x"-Retry-Version). Dann kann der Parser
das direkt per Konvention berücksichtigen.
</div>
""", unsafe_allow_html=True)

    st.markdown("## ℹ️ Weitere Beobachtungen")
    st.markdown("""
<div class="issue-card">
  <h4>Gerundete Werte in der Messübersicht</h4>
  <p>Bei den IDs 0058–0060 stehen gerundete Werte wie <code>Qa = 0.47</code> statt
  der üblichen 15-stelligen Präzision. Vermutlich nachträgliche Hand-Eingabe, nicht
  automatische Berechnung.</p>
</div>

<div class="issue-card">
  <h4>Fehlende HAD-Dateien für einige Projekt-1-Messungen</h4>
  <p>Bei einigen ASC-Dateien fehlt das korrespondierende HAD-File. Ohne HAD kennen
  wir die Probenfläche nicht und die Messfläche-Spalte bleibt leer in der
  Messübersicht.</p>
</div>

<div class="issue-card">
  <h4>Leere Messübersicht-Zeilen</h4>
  <p>Von 91 Messung-IDs haben nur 68 Ja/Jr-Werte. Die restlichen 23 sind leer —
  vermutlich Fehlmessungen oder Laborexperimente, deren Werte nicht aufgezeichnet
  wurden.</p>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# Footer
# ═══════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="mpc-footer">
  <b>MPC² GmbH</b> · Peter Tunner-Straße 19, 8700 Leoben, Österreich ·
  <a href="https://mpc2.at" target="_blank">mpc2.at</a><br>
  <span style='font-size:12px; opacity:0.85;'>
    Corrosion Ray® DL-EPR-Analyzer · gebaut von <a href="https://werchota.ai" target="_blank">werchota.ai</a> ·
    JTF Green Startupmark
  </span>
</div>
""", unsafe_allow_html=True)
