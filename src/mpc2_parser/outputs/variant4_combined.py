"""Variant 4: Combined workbook — per-measurement sheets + summary overview tab.

Single .xlsx containing:
  - Sheet "Übersicht": one row per measurement, ready to copy into master
  - One sheet per measurement (same format as Variant 1)

Best of both worlds — detail when you need it, summary when you don't.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..core import Measurement
from .variant1_auswertung import _add_measurement_sheet


SUMMARY_COLUMNS = [
    "Messung", "Datei", "Material", "Probenbez",
    "Messfläche [mm²]", "Temperatur [°C]", "Probe-T [°C]",
    "Ruhepotential [mV]", "Ja [mA/cm²]", "Jr [mA/cm²]",
    "Jr/Ja", "Qa [As]", "Qr [As]", "Qr/Qa",
    "Split-Index", "Split-Methode", "Datum", "Prüfer", "Bemerkung",
]


def _fill_summary_sheet(ws, measurements: list[Measurement]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F4C5C")

    for col, h in enumerate(SUMMARY_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for r, m in enumerate(measurements, start=2):
        fm = m.filename_meta
        an = m.analysis
        had = m.had

        def safe_round(v, places):
            if v is None or v != v:  # NaN check
                return None
            return round(v, places)

        row_values = [
            fm.messung_id,
            Path(m.source_file).name,
            fm.material,
            fm.probenbez,
            had.probenflaeche_mm2,
            fm.temperature_c,
            fm.probe_temperature_c,
            safe_round(an.ruhepotential_mv, 2),
            safe_round(an.ja_ma_cm2, 5),
            safe_round(an.jr_ma_cm2, 5),
            safe_round(an.jr_ja, 6),
            safe_round(an.qa_as, 8),
            safe_round(an.qr_as, 8),
            safe_round(an.qr_qa, 6),
            an.split_index,
            an.split_method,
            had.erstellungsdatum,
            had.sachbearbeiter,
            fm.notes,
        ]
        for col, val in enumerate(row_values, start=1):
            ws.cell(row=r, column=col, value=val)

    # Column widths
    widths = [10, 45, 14, 10, 14, 14, 14, 16, 14, 14, 10, 14, 14, 10, 12, 14, 12, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C2"


def write_combined_workbook(
    measurements: Iterable[Measurement],
    output_path: Path | str,
    project_name: str | None = None,
) -> Path:
    """Write a combined workbook with summary + per-measurement sheets."""
    output_path = Path(output_path)
    measurements = list(measurements)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. Summary sheet first (default view when file opens)
    summary_ws = wb.create_sheet("Übersicht", 0)
    _fill_summary_sheet(summary_ws, measurements)

    # 2. One sheet per measurement
    for m in measurements:
        try:
            _add_measurement_sheet(wb, m)
        except Exception as e:
            ws = wb.create_sheet(f"ERR_{m.filename_meta.messung_id or '?'}")
            ws["A1"] = f"Error: {e}"

    if project_name:
        wb.properties.title = f"Auswertung {project_name}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
