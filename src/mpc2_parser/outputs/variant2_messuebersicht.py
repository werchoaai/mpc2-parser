"""Variant 2: Append parsed measurements as new rows in the master Messübersicht.

Writes to the "Corrosion Ray" sheet of
    Messübersicht_Elektrochemische Messungen.xlsx

Columns (Corrosion Ray sheet):
  A Messung, B Datum, C Uhrzeit, D Messzelle, E Versuchsnr.,
  F Untersuchungsthema/Order-No., G Material, H Materialzustand,
  I Probenbez, J Messfläche [mm²], K Lösung,
  L H2SO4 [ml/l], M HCl [ml/l], N KSCN [mM/l], O NaCl [mM/l],
  P Messgeschw. [mV/s], Q Vertexpotential [mVSCE], R Vertexpotential [mVAg/AgCl],
  S Temperatur [°C], T RT [°C], U Ruhepotential [mV],
  V Ja [mA/cm²], W Jr [mA/cm²], X Jr/Ja,
  Y Qa, Z Qr, AA Qr/Qa,
  AB Prüfer, AC Bemerkung, AD Zusatz
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl

from ..core import Measurement


COLS = {
    "Messung": "A", "Datum": "B", "Uhrzeit": "C", "Messzelle": "D",
    "Versuchsnr.": "E", "Untersuchungsthema/Order-No.": "F",
    "Material": "G", "Materialzustand": "H", "Probenbez": "I",
    "Messfläche [mm²]": "J", "Lösung": "K",
    "H2SO4 [ml/l]": "L", "HCl [ml/l]": "M", "KSCN [mM/l]": "N", "NaCl [mM/l]": "O",
    "Messgeschw. [mV/s]": "P",
    "Vertexpotential [mVSCE]": "Q", "Vertexpotential [mVAg/AgCl]": "R",
    "Temperatur [°C]": "S", "RT [°C]": "T", "Ruhepotential [mV]": "U",
    "Ja [mA/cm²]": "V", "Jr [mA/cm²]": "W", "Jr/Ja": "X",
    "Qa": "Y", "Qr": "Z", "Qr/Qa": "AA",
    "Prüfer": "AB", "Bemerkung": "AC", "Zusatz": "AD",
}


def _parse_had_date(erstellungsdatum: str | None) -> datetime | None:
    """HAD format: '01.04.2025'"""
    if not erstellungsdatum:
        return None
    try:
        return datetime.strptime(erstellungsdatum.strip(), "%d.%m.%Y")
    except (ValueError, AttributeError):
        return None


def _build_row(m: Measurement, order_no: str | None = None) -> dict:
    """Build a row dict keyed by column letter for the Corrosion Ray sheet."""
    fm = m.filename_meta
    had = m.had
    an = m.analysis

    date = _parse_had_date(had.erstellungsdatum)
    time_str = had.erstellungszeit or None

    return {
        "A": fm.messung_id,
        "B": date,
        "C": time_str,
        "D": "Ray",
        "E": None,  # Versuchsnr. — can't infer from filename reliably
        "F": order_no,
        "G": fm.material,
        "H": None,  # Materialzustand — too variable to parse
        "I": fm.probenbez,
        "J": had.probenflaeche_mm2,
        "K": None,  # Lösung — need solution mix info
        "L": None, "M": None, "N": None, "O": None,
        "P": None,  # Messgeschw.
        "Q": None, "R": None,
        "S": fm.temperature_c,
        "T": fm.probe_temperature_c,
        "U": round(an.ruhepotential_mv, 2) if an.ruhepotential_mv == an.ruhepotential_mv else None,
        "V": round(an.ja_ma_cm2, 5),
        "W": round(an.jr_ma_cm2, 5),
        "X": round(an.jr_ja, 6) if an.ja_ma_cm2 else None,
        "Y": round(an.qa_as, 8),
        "Z": round(an.qr_as, 8),
        "AA": round(an.qr_qa, 6) if an.qa_as else None,
        "AB": had.sachbearbeiter,
        "AC": fm.notes,
        "AD": Path(m.source_file).name,
    }


def append_to_messuebersicht(
    measurements: Iterable[Measurement],
    master_xlsx: Path | str,
    output_xlsx: Path | str | None = None,
    sheet_name: str = "Corrosion Ray",
    order_no: str | None = None,
    overwrite_existing_ids: bool = False,
) -> Path:
    """Append new measurement rows to the master Messübersicht workbook.

    Creates a copy at `output_xlsx` (default: `<master>_updated.xlsx`) to avoid
    overwriting Manuel's original.

    If `overwrite_existing_ids=True`, rows whose Messung ID already exists in
    the sheet are updated in place; otherwise they are skipped with a warning.
    """
    master_xlsx = Path(master_xlsx)
    if output_xlsx is None:
        output_xlsx = master_xlsx.with_name(master_xlsx.stem + "_updated" + master_xlsx.suffix)
    output_xlsx = Path(output_xlsx)

    shutil.copy2(master_xlsx, output_xlsx)
    wb = openpyxl.load_workbook(output_xlsx)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    # Find existing Messung IDs to detect duplicates
    existing = {}
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            existing[str(v).strip()] = r

    appended = 0
    updated = 0
    skipped = 0
    report = []

    for m in measurements:
        row_data = _build_row(m, order_no=order_no)
        mid = str(row_data["A"] or "").strip()

        if mid in existing:
            if overwrite_existing_ids:
                target_row = existing[mid]
                action = "updated"
                updated += 1
            else:
                skipped += 1
                report.append(f"  [SKIP] {mid}: already in sheet (row {existing[mid]})")
                continue
        else:
            target_row = ws.max_row + 1
            action = "added"
            appended += 1

        for col_letter, value in row_data.items():
            if value is not None:
                ws[f"{col_letter}{target_row}"] = value

        report.append(f"  [{action.upper()}] {mid} -> row {target_row}")

    wb.save(output_xlsx)
    print(f"Messübersicht updated: {output_xlsx}")
    print(f"  Appended: {appended}, Updated: {updated}, Skipped: {skipped}")
    for line in report:
        print(line)

    return output_xlsx
