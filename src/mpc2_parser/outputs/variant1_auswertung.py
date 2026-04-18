"""Variant 1: Per-project Auswertung workbook matching Manuel's existing template.

One sheet per measurement with:
  - Columns A-I: raw curve data (with derived columns as formulas)
  - Summary block (L-P): Ja, Jr, Qa, Qr, ratios, Ruhepotential
  - Named identically to Manuel's convention

Drop-in replacement for the current manual workflow.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..core import Measurement


HEADERS = [
    "Zeit [s]",
    "Potenzial [V]",
    "I [A]",
    "J [A/m²]",
    "Potenzial [mV]",
    "J [mA/cm²]",
    "Q [As]",
    "delta t [s]",
    "delta Q [As]",
]


def _sheet_name_from_measurement(m: Measurement) -> str:
    """Build a sheet name matching Manuel's style, e.g. '399_UNS S32906_K2-1'."""
    mid = m.filename_meta.messung_id or "?"
    # strip leading zeros to match Manuel's "399" not "0399"
    mid = mid.lstrip("0") or "0"
    mat = m.filename_meta.material or ""
    probe = m.filename_meta.probenbez or ""
    # Excel sheet names max 31 chars, no []:*?/\
    parts = [p for p in (mid, f"UNS {mat}" if mat else "", probe) if p]
    name = "_".join(parts)
    name = "".join(c for c in name if c not in r"[]:*?/\\")
    return name[:31]


def _add_measurement_sheet(wb: openpyxl.Workbook, m: Measurement) -> None:
    """Add a single measurement as a new sheet in the workbook."""
    name = _sheet_name_from_measurement(m)
    # Handle duplicate names
    base = name
    i = 2
    while name in wb.sheetnames:
        name = f"{base[:28]}_{i}"
        i += 1

    ws = wb.create_sheet(name)

    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8F4F7")
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    a = m.asc
    n = a.n_points
    # Data rows: 2..(n+1). Column A-D are values, E-I are formulas.
    for i in range(n):
        r = i + 2
        ws.cell(row=r, column=1, value=float(a.time_s[i]))
        ws.cell(row=r, column=2, value=float(a.potential_v[i]))
        ws.cell(row=r, column=3, value=float(a.current_a[i]))
        ws.cell(row=r, column=4, value=float(a.current_density_am2[i]))
        ws.cell(row=r, column=5, value=f"=B{r}*1000")          # mV
        ws.cell(row=r, column=6, value=f"=D{r}/10")             # mA/cm²
        ws.cell(row=r, column=7, value=0)                       # Q - starts at 0
        if i < n - 1:
            ws.cell(row=r, column=8, value=f"=A{r+1}-A{r}")     # delta t
            ws.cell(row=r, column=9, value=f"=H{r}*C{r}")       # delta Q
        else:
            ws.cell(row=r, column=8, value=0)
            ws.cell(row=r, column=9, value=0)

    # Summary block using ranges from our analysis
    an = m.analysis
    jr_start, jr_end = an.excel_jr_range
    qa_start, qa_end = an.excel_qa_range
    qr_start, qr_end = an.excel_qr_range

    summary_font = Font(bold=True, color="0F4C5C")

    ws["L2"] = "Ja"
    ws["L2"].font = summary_font
    ws["M2"] = f"=MAX(F{qa_start}:F{qa_end})"  # Ja over forward sweep

    ws["L3"] = "Jr"
    ws["L3"].font = summary_font
    ws["M3"] = f"=MAX(F{jr_start}:F{jr_end})"

    ws["L5"] = "Qa"
    ws["L5"].font = summary_font
    ws["M5"] = f"=SUM(I{qa_start}:I{qa_end})"

    ws["L6"] = "Qr"
    ws["L6"].font = summary_font
    ws["M6"] = f"=SUM(I{qr_start}:I{qr_end})"

    ws["O3"] = "Qr/Qa"
    ws["O3"].font = summary_font
    ws["P3"] = "=M6/M5"

    ws["O4"] = "Jr/Ja"
    ws["O4"].font = summary_font
    ws["P4"] = "=M3/M2"

    ws["O6"] = "Ruhepotential"
    ws["O6"].font = summary_font
    ws["P6"] = "=E87"  # Manuel's convention
    ws["Q6"] = "mV"

    # Info footer
    info_row = 10
    ws.cell(row=info_row, column=12, value="Split-Methode:").font = Font(italic=True)
    ws.cell(row=info_row, column=13, value=an.split_method)
    ws.cell(row=info_row + 1, column=12, value="Split-Index:").font = Font(italic=True)
    ws.cell(row=info_row + 1, column=13, value=an.split_index)
    ws.cell(row=info_row + 2, column=12, value="Reverse-Endpoint:").font = Font(italic=True)
    ws.cell(row=info_row + 2, column=13, value=an.split_diagnostics.get("reverse_endpoint"))
    ws.cell(row=info_row + 3, column=12, value="Datei:").font = Font(italic=True)
    ws.cell(row=info_row + 3, column=13, value=Path(m.source_file).name)

    # Column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(12, 18):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_auswertung_workbook(
    measurements: Iterable[Measurement],
    output_path: Path | str,
    project_name: str | None = None,
) -> Path:
    """Write a per-project Auswertung workbook.

    One sheet per measurement, matching Manuel's template exactly with formulas
    (not hard-coded values) so he can still tweak ranges manually if needed.
    """
    output_path = Path(output_path)
    measurements = list(measurements)
    if not measurements:
        raise ValueError(
            "Keine auswertbaren Messungen vorhanden — "
            "bitte mindestens eine gültige DL-EPR-Datei hochladen."
        )

    wb = openpyxl.Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    count = 0
    for m in measurements:
        try:
            _add_measurement_sheet(wb, m)
            count += 1
        except Exception as e:
            # Add an error sheet so the user sees what failed
            ws = wb.create_sheet(f"ERR_{count+1}")
            ws["A1"] = f"Error processing {Path(m.source_file).name}"
            ws["A2"] = str(e)

    if count == 0:
        # All measurements failed — workbook would have zero visible sheets
        raise ValueError(
            "Alle Messungen konnten nicht verarbeitet werden — "
            "Auswertungs-Workbook wurde nicht erstellt."
        )

    # Optional project-level summary sheet at the front
    if project_name:
        wb.properties.title = f"Auswertung {project_name}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
